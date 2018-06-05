import os,sys
import glob
from openpyxl import load_workbook
from openpyxl.compat import range

Excelfilename = 'excel_combined.xlsx'
# Open existing file with load_workbook
wb = load_workbook(Excelfilename)
# Create a new sheet instead of renaming active
ws = wb.create_sheet('Sheet1')

read_files = glob.glob("split_speed/*.txt")
i = 0
for file in read_files:
	base=os.path.basename(file)
	basename=os.path.splitext(base)[0]
	# sheet_name = "excel_new/"+basename+".xls"
	# # print(basename,",", sheet_name)
	# workbook = xlwt.Workbook(encoding="utf-8")
	# sheet1 = workbook.add_sheet("Sheet1", cell_overwrite_ok=True)
	# b=1
	# count=1
	# while b<=96:
	# 	a=0
	# 	if (b-1)%4 == 0:
	# 		time_n=(b-1)/4
	# 		time_compare = 0 
	# 	elif (b-2)%4 == 0:
	# 		time_n=(b-2)/4
	# 		time_compare = 15
	# 	elif (b-3)%4 == 0:
	# 		time_n=(b-3)/4
	# 		time_compare =30
	# 	else:
	# 		time_n=(b-4)/4
	# 		time_compare = 60
	# 	if b<=96:
	# 		time_n = '%02d' % time_n
	# 		time_compare = '%04d' % time_compare
	# 		data = str(time_n) + str(time_compare)
	# 		print(data)
	# 		# count+=1
	# 		sheet1.write(a,b-1,data)
	# 	b+=1
	# print(count)
	i+=1
	ws.cell(column=1, row=i, value=basename)
	with open(file,encoding="utf-8") as myFile:
	    for line in myFile:
	        date = int(line[6:][:2]) ##change basename to line
	        # print(date)
	        time=line[9:][:6]
	        # print(time)
	        # temp = 0
	        # i=date
	        j=0
	        # while temp<=31:
	        #     temp+=1
	        #     if temp == date:
	        #         i = date
	        #         print(i,"\n\n\n")
	        time_n = int(time[:2])
	        # print(time_n)
	        time_compare = int(time[2:])
	        # print(time_compare)
	        if 0<=time_compare and time_compare<1500:
	            j = 4*time_n + 1 +(date-1)*96
	        elif 1500<=time_compare and time_compare<3000:
	            j = 4*time_n + 2 +(date-1)*96
	        elif 3000<=time_compare and time_compare<4500:
	            j = 4*time_n + 3 +(date-1)*96
	        else:
	            j = 4*time_n + 4 +(date-1)*96
	        # print(j,"\n\n\n")
	        line=line[:-1]
	        speed_data = line[16:] ##check for the data once
	        # print(speed_data)
	        # sys.stdout = open("filename.txt","a")
	        # print(i,",",j,"\n",end="",sep="")
	        # sys.stdout.close()
	        ws.cell(column=j+1, row=i, value=speed_data)
	    wb.save(filename = Excelfilename)
